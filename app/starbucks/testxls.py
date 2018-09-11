from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter,column_index_from_string


rd_file ='testpy.xlsx'
out_file="testcase.xlsx"
name_col=0
type_col=2
mandatory_col=4
check_col=5
example_col=6

urlcell ="B1"
methodcell ="B2"

data_row = 4


def sortbytype(elem):
	return elem[type_col]


def format_to_row(header,dct):
	formdata = ["" for i in range(len(header))]
	#print(header)
	for k,v in dct.items():
		if k in header:
			formdata[header.index(k)]=v
	return tuple(formdata)


def get_col(data,col_num):
	return [j[col_num] for j in data]

def dict_two_col(data,key_num,val_num,filter_col=False):
	dct =dict(zip(get_col(data,key_num),get_col(data,val_num)))
	if filter_col == False:
		return dct
	elif filter_col == True:
		return {k: v for k, v in dct.items() if v !=None and str(v).strip()!=""}


def creat_case_row(example_dct,target_dct,repl,casetext,expecttext,outdata,header):
	for key in target_dct:
		kong =example_dct.copy()
		kong[key] =repl
		kong["Case ID"]= casetext % key
		kong["Expected Results"]= expecttext % key
		outdata.append(format_to_row(header,kong))


def write_row(ws,row,column,iter):
	col =column
	for i in iter:
		ws.cell(row=row, column=col).value = i 
		col=col+1






class Testcase(object):
	def __init__(self):
		self.wb =load_workbook(rd_file)
		self.sheetnames = list(self.wb.sheetnames)[:1]#测试一个sheet


	def parse_one_sheet(self,sheetname):
		#print(sheetname)
		sheet = self.wb[sheetname]
		url =sheet[urlcell].value
		method =sheet[methodcell].value
		data =[]

		# get the talle into data
		for row in list(sheet.rows)[data_row:]:
			line = tuple([col.value for col in row])[:example_col+1]
			if line[0]!=None:
				data.append(line)

		data.sort(key=sortbytype,reverse=True)
		#print(data)

		##return deader 
		name_param =get_col(data,name_col)
		header =["Case ID",*name_param,"Expected Results","Actual Results"]
		
		header_dct =dict_two_col(data,name_col,type_col)


		### check type
		type_param =get_col(data,type_col)
		print(type_param)
		for i in type_param:
			if i not in ("header","form"):
				raise Exception("type in the excel is not header or form")


		
		outdata =[]

		#print(dict_two_col(data,name_col,example_col,True))
		example_dct = dict_two_col(data,name_col,example_col,True)


		## check mandatory col
		mandatory_param = get_col(data,mandatory_col)
		for ck in mandatory_param:
			if ck not in (True,False):
				raise Exception("please check the mandatory")



		### header row
		#outdata.append(tuple(header))

		####### first row,accroding mandatory and  example row
		example_dct["Case ID"]= "all_valid_value"
		example_dct["Expected Results"]= "HTTP Code 200"
		outdata.append(format_to_row(header,example_dct))


		####### mandatory case ,check if the mandatory must be true and replace with '-'
		mandatory_dct =dict_two_col(data,name_col,mandatory_col)
		mandatoryitems ={k: v for k, v in mandatory_dct.items() if v ==True}
		creat_case_row(example_dct,mandatoryitems,"-","mandatory param %s not given","param (%s) must be gived.",outdata,header)

		'''
		for key in mandatoryitems:
			kong =example_dct.copy()
			kong[key] ='-'
			kong["Case ID"]= "mandatory param %s not given" % key
			kong["Expected Results"]= "param (%s) must be gived." % key
			outdata.append(format_to_row(header,kong))
		'''


		####### empty and vaild check .
		check_dct =dict_two_col(data,name_col,check_col,True)
		emtpy_ck ={k: v for k, v in check_dct.items() if 'empty' in v }
		creat_case_row(example_dct,emtpy_ck,"","empty param %s while others vaild","param (%s) must not be empty.",outdata,header)

		valid_ck={k: v for k, v in check_dct.items() if "invalid" in v}
		creat_case_row(example_dct,emtpy_ck,"","invalid param %s while others vaild","param (%s) is not right.",outdata,header)

		return url,method,header,outdata,header_dct

		### write to excel
		#write_to_excel(url,method,outdata,out_file)




	def format_to_excel(self,wb,pagenum,sheetname):
		url,method,header,outdata,header_dct =self.parse_one_sheet(sheetname)

		title =url.split('/')[-1]
		ws = wb.create_sheet("1."+str(pagenum)+" "+title)



		title_cell = ws['B3']
		title_cell.value = "1."+str(pagenum)+" "+title
		ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=len(header)+2)



		ws['B4'].value= "Type"
		ws['B5'].value= "Method"
		ws['B6'].value= "Header"
		ws['B7'].value= "Url"

		ws['C4'].value= "RESTful JSON API."
		ws['C5'].value= method
		ws['C6'].value= "Accept:application/json"
		ws['C7'].value= url
		
		ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=len(header)+2)
		ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=len(header)+2)
		ws.merge_cells(start_row=6, start_column=3, end_row=6, end_column=len(header)+2)
		ws.merge_cells(start_row=7, start_column=3, end_row=7, end_column=len(header)+2)

		ws['B8'].value= "Priority:"

		header2 = header[1:-2] 
		#print(header2)
		ws['C8'].value= "Case ID"

		#header_dct_fm= self.header_repl(method,header_dct)
		write_row(ws,8,4,list(header_dct.values()))
		write_row(ws,8,4+len(header2),["Expected Results","Actual Results"])
		write_row(ws,9,4,list(header_dct.keys()))

		#start from 10

		start =10
		for rowdata in outdata:
			write_row(ws,start,3,rowdata)
			start+=1

		self.format_tyle(ws)
		

	def format_tyle(self,ws):
		'''
		3-7行 行高25
		10-max 行高 35
		'''
		max_row =ws.max_row
		ws.row_dimensions[3].height =34
		for i in range(4,8):
			ws.row_dimensions[i].height =25
		for i in range(10,max_row+1):
			ws.row_dimensions[i].height =35

		'''
		第二列 12
		第三caseID列 50
		其他列 循环 20

		导数第二列 50
		导数第一列 20

		'''

		max_col=ws.max_column
		#print(max_col)
		ws.column_dimensions['A'].width= 4
		ws.column_dimensions['B'].width= 12
		ws.column_dimensions['C'].width= 50

		for i in range(4,max_col-2+1):
			ws.column_dimensions[get_column_letter(i)].width= 20

		ws.column_dimensions[get_column_letter(max_col-2+1)].width=50
		ws.column_dimensions[get_column_letter(max_col)].width=20



		#style
		f18_bold =Font(size=18,bold=True)
		center= Alignment(horizontal="center",vertical="center")

		#border
		#thick = Side(border_style="thick", color="000000")
		thin = Side(border_style="thin", color="000000")
		border = Border(top=thin, left=thin, right=thin, bottom=thin)

		# fill
		bluefill = PatternFill("solid", fgColor="8DB4E2")
		lightyellow =PatternFill("solid", fgColor="FABF8F")
		darkyellow=PatternFill("solid", fgColor="FFC000")


		
		ws["B3"].font =f18_bold
		ws["B4"].font =f18_bold
		ws["B5"].font =f18_bold
		ws["B6"].font =f18_bold
		ws["B7"].font =f18_bold

		ws["B4"].fill =bluefill
		ws["B5"].fill =bluefill
		ws["B6"].fill =bluefill
		ws["B7"].fill =bluefill
		ws["C4"].fill =bluefill
		ws["C5"].fill =bluefill
		ws["C6"].fill =bluefill
		ws["C7"].fill =bluefill

		for row in ws['C8':get_column_letter(ws.max_column)+str(8)]:
			for cl in row:
				cl.fill =lightyellow
		for row in ws['C9':get_column_letter(ws.max_column-2)+str(9)]:
			for cl in row:
				cl.fill =darkyellow





		ws["B3"].alignment =center

		
		ws["B4"].border =border

		for row in ws['B3':get_column_letter(ws.max_column)+str(ws.max_row)]:
			for cellit in row:
				cellit.border =border
		







	def main(self):
		pagenum =0
		wb = Workbook()

		
		for sheetname in self.sheetnames:
			pagenum +=1

			self.format_to_excel(wb,pagenum,sheetname)
			#url,method,header,outdata =self.parse_one_sheet(sheetname)

			#title =url.split('/')[-1]
			#ws = wb.create_sheet("1."+str(pagenum)+" "+title)
		wb.save('testcase.xlsx')
			





#for row in list(sheet.rows)[4:]:
#	line = [col.value for col in row]
#	print(line)


testcase =Testcase()
testcase.main()


